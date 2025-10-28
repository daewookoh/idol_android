#!/usr/bin/python
# -*- coding:utf-8 -*-
#
# dependencies
#   openpyxl==2.3.0
#
# 1st row must be like this:
#  KEY | default | ko | es | ...

import os
import shutil
import sys
import codecs
from openpyxl import load_workbook
import argparse

# android:ios table
language_table = {"zh-rcn":"zh-Hans", "zh-rtw":"zh-Hant", "in":"id", "pt":"pt-BR"}
OUTPUT_DIR = 'output'
TAB = ' ' * 4

parser = argparse.ArgumentParser(description='Generate strings.xml files from xlsx file.')
parser.add_argument("-i", metavar="<file>", help="import from xlsx file")
parser.add_argument("-o", metavar="<output directory>")
args = parser.parse_args()
if args.i:
    file = args.i
else:
    print 'Usage: \n    import: langtool.py -i xlsx_filename'
    sys.exit()
    
print '***** DEPRECATED: USE google_langtool.py INSTEAD. *****'
    
wb = load_workbook(filename = file)
ws = wb.get_active_sheet()

col = 0

if os.path.isdir(OUTPUT_DIR):
    shutil.rmtree(OUTPUT_DIR)
os.makedirs(OUTPUT_DIR)

while not ws[chr(ord('C') + col) + '1'].value is None:
    # read language code
    lang_code = ws[chr(ord('C') + col) + '1'].value
    
    print 'processing ' + lang_code

    # android    
    if lang_code == 'default':
        folder = OUTPUT_DIR + '/android/values'
    else:
        folder = OUTPUT_DIR + '/android/values-' + lang_code
    
    os.makedirs(folder)
    f = codecs.open(folder + '/strings.xml', 'w', encoding='utf-8')
    f.write('<?xml version="1.0" encoding="utf-8"?>\n<resources>\n')

    # ios
    if lang_code == 'default':
        folder_ios = OUTPUT_DIR + '/ios/Base.lproj'
    else:
        # convert android language code to ios language code
        if lang_code in language_table:
            lang_code = language_table[lang_code]
        
        folder_ios = OUTPUT_DIR + '/ios/' + lang_code + '.lproj'
    
    os.makedirs(folder_ios)
    f_ios = codecs.open(folder_ios + '/Localizable.strings', 'w', encoding='utf-8')
    
    line = 2
    is_empty = False

    # Loop while there is a key    
    while not (ws['A' + str(line)].value is None):
        # read key
        key = ws['A' + str(line)].value
        value = ws[chr(ord('C') + col) + str(line)].value
        
#         if key != None and value != None:
#             print key + ':' + value
                
        # comment ?
        if key[:2] == '/*' or key[:2] == '<!' :
            f.write(TAB + key + '\n')
            f_ios.write( key.replace('<!--', '/*').replace('-->', '*/') + '\n')
        else:
            if value is None:
                value = ''
            # check translatable (android only)
            translatable = ''
            if ws['B' + str(line)].value == False:
                # translatable false
                translatable = ' translatable="false"'

            # write to file is translatable or first column (default language)
            # skip if android & translation is empty
            if (translatable == '' or col == 0) and value != '*' :
                # escape apostrophe
                value_android = value.replace("'", "\\'").replace("\\\\'", "\\'")
                
                string = ( TAB + '<string name="%s"' + translatable + '>%s</string>\n' ) % ( key, value_android )
                f.write( string )
            
            # iOS : fill same string if NOT translatable
            if translatable != '' or value == '*' :
                value = ws['C' + str(line)].value
            # unescape
            if  value is not None:
                value_ios = value.replace("&gt;", ">").replace("&lt;", "<").replace("\\'", "'").replace('"', '\\"').replace('\\\\"', '\\"').replace('&amp;', '&')
                string = ( '"%s" = "%s";\n' ) % ( key, value_ios )
                f_ios.write( string )
            
        line += 1
        
    f.write('</resources>')
    f.close()
    
    f_ios.close()
    
    print 'done.'
    print ''    
    
    col += 1
    
# copy Base.lproj to en.lproj
shutil.copytree( OUTPUT_DIR + '/ios/Base.lproj', OUTPUT_DIR + '/ios/en.lproj' )
